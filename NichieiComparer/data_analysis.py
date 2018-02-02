"""
Access to analysis on documents translated and untranslated.
The connection between natto and MeCab is very finicky,
and MeCab will often fail to initialize.
"""
import os
import logging
from re import search
from collections import namedtuple, Counter
from openpyxl import load_workbook, Workbook
from natto import MeCab, MeCabError
from NichieiComparer.utils import FileIO, LangUtils, ExcelUtils, CollectionUtils

KeyTuple = namedtuple('KeyTuple', ['root_form', 'katakana']) # pylint: disable=invalid-name
log = logging.getLogger('NichieiComparer.data_analysis') # pylint: disable=invalid-name
log.setLevel(logging.DEBUG)

class ExcelDoc:
    """An xlsx document that requires or contains a translation from Japanese
    to some other language, English by default."""

    def __init__(self, filepath: str, translated: bool):
        """Automatically opens a workbook and analyzes it or opens a previously generated
        analysis JSON record."""
        assert os.path.exists(filepath)
        self._doc_path = filepath
        self._json_path = os.path.splitext(filepath)[0] + ".json"
        self.pairs = [] # List of source / translation tuples in order of appearance
        self.terms = {} # KeyTuple -> List of pair indices where the term comes up
        # First check to see if it already has been analyzed
        # TODO: Update json when original file changes
        if not self._load_analysis():
            workbook = load_workbook(self._doc_path)
            if translated:
                self.pairs = ExcelDoc._load_pairs_from_workbook(workbook) # Get pairs
            else:
                self.pairs = ExcelDoc._find_source_in_workbook(workbook) # Get half pairs (source)
            self.terms = ExcelDoc._get_source_terms(self.pairs)
            self._save_analysis()
                
    def _save_analysis(self):
        """Hides the process by which a JSON analysis file is saved."""
        log.debug("Saving analysis JSON...")
        json_dict = {"pairs": self.pairs, "terms": {}}
        for kterms, rows in self.terms.items():
            json_dict["terms"]["[{}][{}]".format(*kterms)] = rows
        FileIO.save_json_dict(self._json_path, json_dict)

    def _load_analysis(self):
        """Hides the process by which a JSON analysis file is loaded."""
        if os.path.exists(self._json_path):
            try:
                log.debug("Loading analysis JSON for %s", os.path.split(self._doc_path)[1])
                # Load previously analyzed data
                json_dict = FileIO.get_json_dict(self._json_path)
                self.pairs = json_dict["pairs"]
                for unparsed_key, row_list in json_dict["terms"].items():
                    # The keys are saved as strings: [root form][katakana pronunciation used]
                    assert unparsed_key[0] == "[" and unparsed_key[-1] == "]"
                    key = search(r'^\[(.*?)\]\[(.*?)\]$',unparsed_key).groups()
                    self.terms[key] = row_list
                return True
            except FileNotFoundError:
                pass
        return False        

    @staticmethod
    def _find_source_in_workbook(wkbk: Workbook):
        """Extracts Japanese text from a workbook. Use on untranslated workbooks.
        Columns are identified as source material by the number of consecutive cells 
        (consistently adjacent) holding Japanese text"""

        alt_japanese = []

        for sheet in wkbk:
            jpn_col_cells = {}
            jpn_col_cell_total = {}
            for row_num, row in enumerate(sheet.iter_rows(), start=1):
                for col_num, cell in enumerate(row, start=1):
                    if LangUtils.is_japanese(cell.value):
                        jpn_col_cells[col_num] = jpn_col_cells.get(col_num, []) + [row_num]

            # Identify and remove potential "header" rows
            log.debug("Removing rows that contain Japanese text in nearly all columns...")
            all_rows = CollectionUtils.flatten(jpn_col_cells.values())
            # Count the number of times Japanese text appears in a row
            jpn_in_row = Counter(all_rows)
            for row_num, count in jpn_in_row.most_common():
                if count > len(jpn_col_cells.keys()) * 0.5:
                    for row in jpn_col_cells.values():
                        row.remove(row_num)

            # Using the columns_with_japanese set,
            # find a column with many consecutive values
            for col, rows in jpn_col_cells.items():
                jpn_col_cell_total[col] = CollectionUtils.total_adjacent_values(rows)
                          
            # Sort the items of jpn_col_cell_total to find the column with the most cells
            # sorted_cols: (col num, consecutive values) in decreasing order of consecutive values
            sorted_cols = [(x, y) for x, y in reversed(
                sorted(jpn_col_cell_total.items(), key=lambda i: i[1])
            )]
            most_likely_column = sorted_cols[0][0]
            mlc_char = ExcelUtils.col_cipher(most_likely_column)
            mlc_good = False

            if len(sorted_cols) >= 3:
                if sorted_cols[0][1] > sorted_cols[1][1] + sorted_cols[2][1]:
                    mlc_good = True
            else:
                mlc_good = True

            if mlc_good:
                alt_japanese += [
                    (sheet["{}{}".format(mlc_char, row_num)].value, "") for
                    row_num in jpn_col_cells[most_likely_column]
                ]

        return alt_japanese

    @staticmethod
    def _load_pairs_from_workbook(wkbk: Workbook):
        """Extracts pairs of source (Japanese) and translation (English) text.
        """
        alt_japanese = []

        for sheet in wkbk:
            # Get sheet dimensions
            # dim_str = sheet.calculate_dimension()
            # start_col, start_row, end_col, end_row = re.search(
            #     "([A-Z]+)([0-9]+):([A-Z]+)([0-9]+)", dim_str).groups()
            alt_col_cells = {} # Dictionary of Cells that have switched to or from Japanese text
                            # Columns where cells switched to or from Japanese text
                            # Key: Column number -> (row number, is japanese percentage)
            #cell_text_lengths = {}
            alt_col_max_cell_count = 0
            # Find likely source text column and start row by finding an
            # alternation between Japanese and non-Japanese on the same row
            # Create meta data to understand the number of cells in each
            # column actually holding data, and their language
            for row_num, row in enumerate(sheet.iter_rows(), start=1):
                prev_japanese = None
                for col_num, cell in enumerate(row, start=1):
                    #cell_text_lengths[(col_num, row_num)] = len(str(cell.value or ""))
                    cell_is_japanese = LangUtils.is_japanese(cell.value)
                    if(
                        prev_japanese is not None and
                        cell_is_japanese is not prev_japanese and
                        cell_is_japanese is not None
                    ):
                        # Cell has alternated between languages (presumably)
                        alt_col_cells[col_num] = alt_col_cells.get(col_num, [])
                        alt_col_cells[col_num].append((row_num, cell_is_japanese))
                        if len(alt_col_cells[col_num]) > alt_col_max_cell_count:
                            alt_col_max_cell_count = len(alt_col_cells[col_num])
                    prev_japanese = cell_is_japanese

            # Remove columns with filled cells at less that 50% of max (such as comment columns)
            # col_usable_cells = {col_num: len(cell_list) for col_num, cell_list in alt_col_cells.items()}
            # max_usable_cells = max(col_usable_cells.values())
            del_few_cells = [col_num for col_num, cell_list in alt_col_cells.items() if len(cell_list) < alt_col_max_cell_count * 0.5]
            for col_num in del_few_cells:
                del alt_col_cells[col_num]

            assert len(alt_col_cells.keys()) == 2
            col_a_num = min(alt_col_cells.keys())
            col_b_num = max(alt_col_cells.keys())
            assert col_a_num != col_b_num
            col_a_is_jpn_text = sum([is_jpn_text for row_num, is_jpn_text in alt_col_cells[col_a_num]]) > len(alt_col_cells[col_a_num]) * 0.5
            col_b_is_jpn_text = sum([is_jpn_text for row_num, is_jpn_text in alt_col_cells[col_b_num]]) > len(alt_col_cells[col_b_num]) * 0.5
            assert col_a_is_jpn_text and not col_b_is_jpn_text
            col_a_addr = ExcelUtils.col_cipher(col_a_num)
            col_b_addr = ExcelUtils.col_cipher(col_b_num)

            # Now that we're confident that these columns are, in fact,
            # the "Japanese -> Target Language" columns, we should line up
            # the Japanese text with the target language text and return it.
            pairs = []
            for row_data in alt_col_cells[col_a_num]:
                pairs.append((sheet["{}{}".format(col_a_addr, row_data[0])].value, sheet["{}{}".format(col_b_addr, row_data[0])].value))

            alt_japanese = alt_japanese + pairs

            # Use standard deviation to find likely columns
            # alt_col_stdevs = {}
            # for alt_col_num, row_list in alt_col_cells.items():
            #     for row_num, is_japanese in row_list:
                    

            # for col_num in range(1, col_cipher(end_col)):
            #     row_lengths = [lengths[(col_num, x)] for x in range(1, int(end_row))]
            #     stdevs.append(stdev(row_lengths))

            # col_a, col_b = n_highest_indices(stdevs, 2)
            # assert col_a in alt_cols and col_b in alt_cols
            """
            start_row = None # Reuse start_row to show where the actual source/translation begins
            end_row = None # Reuse start_row to show where the alternating text ends
            for alt_cell in alt_cells:
                if alt_cell[1] == col_a or alt_cell[1] == col_b:
                    start_row = start_row or alt_cell[0]
                    end_row = alt_cell[0]

            col_a_val = sheet.cell(row=start_row, column=col_a).value
            col_b_val = sheet.cell(row=start_row, column=col_b).value

            if is_japanese(col_a_val) and not is_japanese(col_b_val):
            """

        return alt_japanese

    @staticmethod
    def _get_source_terms(pairs, retries=10):
        """Takes a list of Japanese->target language pairs and creates a dictionary
        with keys being tokenized Japanese and the values being a list of
        translated sentences that are assumed to have the term inside them."""
        term_dict = {}
        try:
            with MeCab() as mecab:
                for row_num, pair in enumerate(pairs):
                    # Parse Japanese half of pair into nodes
                    for mecab_node in mecab.parse(pair[0], as_nodes=True):
                        if mecab_node.is_nor():
                            features = mecab_node.feature.split(',')
                            if LangUtils.is_useful_term_jp(features):
                                # 6=root, 8=actual pronunciation (last element)
                                key = KeyTuple(features[6], features[8])
                                term_dict[key] = term_dict.get(key, []) + [row_num]
        except MeCabError as init_error:
            if retries > 0:
                try_text = "tries"
                if retries == 1:
                    try_text = "try"
                log.warning("MeCab failed to init (%d %s remaining)", retries, try_text)
                return ExcelDoc._get_source_terms(pairs, retries-1)
            else:
                raise init_error
        return term_dict
